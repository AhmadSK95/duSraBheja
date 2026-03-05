"""Excel/XLSX extraction via openpyxl."""

import asyncio


async def extract_excel(file_path: str) -> str:
    """Extract text content from Excel file."""

    def _extract(path: str) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))

        wb.close()
        return "\n".join(lines)

    return await asyncio.get_event_loop().run_in_executor(None, _extract, file_path)
